# pylint: disable=line-too-long, C0103
#Curve_3.1

import glob
import os

import sys
import threading
from tkinter import (Button, Entry, Frame, IntVar, Label, LabelFrame, Listbox,
                     Menu, Scrollbar, StringVar, Toplevel,
                     filedialog, ttk)
from tkinter.constants import *


import lasio as LS  # pip install lasio

import numpy as np
import pandas as pd  # For Excel actions
from openpyxl import load_workbook
from openpyxl.cell import _writer  # for py installer.

import classes

import global_vars
from classes.Login import Login
from classes.Project import Project
from classes.ExcelEditor import ExcelEditor

from defs import alert, common, excel, las, main, prompt, tools
from enumerables import Dir, ErrorMessage 

#==========================================================================================================
# Exit CURVE
def c_Exit():
    if not prompt.yesno('Are you sure you wish to Quit (Y/N)?'):
        return
    # Save files
    global_vars.project.SaveInit()
    global_vars.project.currentWellList.SaveAs('init.wnz')
    global_vars.project.currentWellList.SaveAs('prev.wnz')
    global_vars.project.Close()
    print('Gone to Exit')
    global_vars.ui.Root.window.quit()

    global_vars.running = False

    # Check for any threads still running
    print("Active threads:")
    for thread in threading.enumerate():
        print(thread.name, thread.is_alive())


    print(' exit')
    quit()




#===================================================================================================================
#----------------------------------------------------------------------------------------------------------------------
# Report Erorrs
# ----------------------------------------------------------------------------------------------------------
def c_ListDir():
    '''
    List Project Directories
    '''

    #create directory list
    mydirlist=[
            'Raw Data                                  : '+global_vars.project.rawDir,
            'Input Data                                : '+global_vars.project.inDir,
            'Calculations                             : '+global_vars.project.outDir,
            'Core Data                                 : '+global_vars.project.coreDir,
            'Cient data                                : '+global_vars.project.clientDir,
            f'Mnemonics, Aliases, Initfiles: {os.getcwd()}',
            'list & plot files                         : '+global_vars.project.inDir+'/databases']

    # Set List Dir toplevel window
    ld_win = common.initCommonWindow(title='Current Project Directories', width=700, height=200, topMost=2)

    mw_names=StringVar(value=mydirlist)
    height=len(mydirlist)
    width=int(200)
    mdir_box=Listbox(ld_win,listvariable=mw_names, height=height, width=width)
    mdir_box.columnconfigure([0,1],weight=1)
    mdir_box.config(selectmode=MULTIPLE)   # select one or more wells
    mdir_box.pack()

    # done
    b_done=Button(ld_win, bg='cyan', text="Done", padx=10, command= lambda : ld_win.destroy())
    b_done.pack(pady=20)
#=====================d===================================================================================
#Load lithology data - excel
def ld_c_des():
    '''
    Load core description excel
    '''

    #locate the spreadsheet in the project directory

    try:
        core = filedialog.askopenfilename(
            initialdir = global_vars.project.coreDir,
            title="Please, select an core description file",
            filetype=(("Excel files", "*.xlsm"), ("all files", "*.*"))
            )
    except FileNotFoundError:
        alert.Error('Directory could not be found. Please, try again.')
    except ValueError:
        alert.Error('File could not be opened. Please, try again.')
    if core=='':                     #canceled
        return

    wb=load_workbook(core,data_only=True)          #open an Excel workbook core description
    ws=wb['Detailed']                  #open the Detailed sheet
    #collect pertinent data

    #Prepare new cas file
    UWI=ws['B5'].value
    path=global_vars.project.inDir+'/'+UWI+'.las'
    new_cas=tools.create_cas(path)      #.cas Core ASCII file format similar to .las
    if not new_cas:   #las cleanup failed
        return
    #new_df = new_cas.df()
    UWI=ws['B5'].value
    #Geologist=ws['H6'].value
    #Datedescr=ws['H7'].value
    #KB=ws['B6'].value
    #Coringfl=ws['H8'].value
    ctop=ws['B10'].value
    new_cas.well.STRT=ctop
    cbase=ws['B13'].value
    new_cas.well.STOP=cbase
    cstep=ws['B14'].value
    new_cas.well.STEP=cstep
    PHIshale=ws['AW14'].value
    Carbshale=ws['AW16'].value
    Pyrshale=ws['AU16'].value
    Sidshale=ws['AU14'].value
    VCLshale=ws['AU18'].value
    new_cas.params['GEO'] = LS.HeaderItem('GEO', value=ws['H6'].value, descr="Core examining geologist")
    new_cas.params['EKB'] = LS.HeaderItem('EKB', unit='M', value=ws['B6'].value)
    new_cas.params['DSD'] = LS.HeaderItem('DSD', value=ws['H7'].value, descr='Date of core description')
    new_cas.params['CFL'] = LS.HeaderItem('CFL', value=ws['H8'].value, descr='Coring fluid')
    new_cas.params['CSH'] = LS.HeaderItem('CSH', unit='M', value=0.00, descr='Core depth shift')

    minrow=35
    maxrow=minrow+int((cbase-ctop)/cstep)
    #determine last depth cell in ws
    addr='A'+str(maxrow)
    if float(ws[addr].value)>cbase:
        addr='A'+str(maxrow-1)
    elif float(ws[addr].value)<cbase:
        addr='A'+str(maxrow+1)
    list_len=maxrow-minrow+1

    #iterate through data columns and convert to dataframe series and attach to core_df
    c_idx=0
    core_no=1
    mlist=[]    #list of curves
    mlist.append([]) #add first curve

    #get number of curve
    mend=len(global_vars.cdescrvs)
    #set null
    mynull=round(new_cas.well.NULL.value,2)
    #columns before striplog
    for col in ws.iter_cols(min_row=minrow, max_row=maxrow,min_col=1,max_col=4, values_only=True):

        for cell in col:
            if c_idx==0:                             #1st Column
                mlist[c_idx].append(round(cell,2))
            if c_idx==1:                             #2nd Column
                if cell==None or cell=='#VALUE!':
                    mlist[c_idx].append(mynull)
                else:
                    mlist[c_idx].append(round(cell,4))
                if isinstance(mlist[c_idx][-1], float):
                    if round(mlist[c_idx][len(mlist[c_idx])-1],4)==0.0:
                        core_no +=1  # next core
            if c_idx==2:                            #3rd Column
                if cell==None or cell=='#VALUE!':
                    mlist[c_idx].append(mynull)
                else:
                    mlist[c_idx].append(cell)
            if c_idx==3:                            #4th Column
                if cell==None or cell=='#VALUE!':
                    mlist[c_idx].append(mynull)
                else:
                    mlist[c_idx].append(round(cell,4))
        c_idx +=1                            #add next curve
        mlist.append([])

    #Now load upto remarks column
    for col in ws.iter_cols(min_row=minrow, max_row=maxrow,min_col=33,max_col=52, values_only=True):
        for cell in col:
            if c_idx>3:                                         #later Columns
                if cell==None or cell=='#VALUE!'or cell==' 'or cell=='':
                    mlist[c_idx].append(mynull)
                elif isinstance(cell, str):
                    mlist[c_idx].append(cell)
                else:
                    if c_idx>9 and c_idx<21:        #Convert PU to fractions
                        mlist[c_idx].append(round(cell/100,4))
                    else:
                        mlist[c_idx].append(round(cell,4))

        c_idx +=1
        mlist.append([])                                        #add next curve

    #Now load after remarks column
    for col in ws.iter_cols(min_row=minrow, max_row=maxrow,min_col=54,max_col=64, values_only=True):
        for cell in col:
            if c_idx>3 and c_idx<35:                   #last Columns
                if cell==None or cell=='#VALUE!':
                    mlist[c_idx].append(mynull)
                else:
                    mlist[c_idx].append(round(cell/100,4))
        if c_idx<35:
            mlist.append([])
        c_idx +=1
        midx=c_idx
        
    # Add empty core analysis curves
    for idx in range(mend-midx+1):
        for l_idx in range(list_len):
            mlist[c_idx].append(mynull)
        if c_idx<mend:
            mlist.append([])
        c_idx +=1

    #Create core_df
    global_vars.core_df=pd.DataFrame(columns=global_vars.cdescrvs,index=mlist[0])   #CORE_DF IS GLOBAL
    #append the data
    c_idx=1
    for crv in mlist[1:-2]:
        global_vars.core_df[global_vars.cdescrvs[c_idx]]=crv
        c_idx +=1

    #mynull=new_cas.well.NULL.value    DOESN'T WORK BECAUSE THERE IS NO DF IN new_cas yet
    #core_df.fillna(mynull)
    #new_cas.set_data(core_df)

    #Create CAS or Core ASCII file
    #First create the DEPT curve
    for c_idx in range(35):
        #mnemonic=cdescrvs[c_idx]
        new_cas.insert_curve(ix=c_idx, mnemonic=global_vars.cdescrvs[c_idx], data=np.array(mlist[c_idx]),
                            unit= global_vars.c_units[c_idx], descr=global_vars.c_descr[c_idx],
                            value='')

    # ready to save cas file
    filename=UWI+'.cas'
    las.save_cas(filename,new_cas, Dir.In)
#-------------------------------------------------------------------------------------------------------------
def ld_c_an():
    '''
    Load Core Analysis from Excel summary excel
    '''

    #locate the spreadsheet in the project directory

    #locate the spreadsheet in the project directory

    try:
        core_an = filedialog.askopenfilename(
            initialdir = global_vars.project.inDir + '/databases',
            title="Please, select a core analysis file",
            filetype=(("Excel files", "*.xlsx"), ("all files", "*.*"))
            )
    except FileNotFoundError:
        alert.Error('Directory could not be found. Please, try again.')
        return
    except ValueError:
        alert.Error('File could not be opened. Please, try again.')
        return

    try:
        wb=load_workbook(core_an,data_only=True)          #open an Excel workbook core analysis data
        ws=wb.active                  #open the Detailed sheet
    except PermissionError:
        alert.Error('File opened elsewhere. Please, close file first.')
        return

    #Create list of UWIs in core analysis database (Excel Work Book)
    idx=0
    U_list=[]
    U_list.append(global_vars.project.currentWellList[0])
    #Column A from row 1 to end
    for cell in ws['A'][1:]:
        #check if cell.value in Well_list i.e. it has a regular las file
        if cell.value in global_vars.project.currentWellList:
            if U_list[idx]!=cell.value:
                idx +=1
                U_list.append(cell.value)
            #else skip

    # Create filelist
    curDir = os.getcwd()
    os.chdir(global_vars.project.inDir)

    # cas_list is the total project in the Input directory
    cas_list = glob.glob("*.cas")

    os.chdir(curDir)                    #restore working directory

    #if there is no CAS file for the UWI it must be created
    #else load cas file
    for UWI in U_list:  #check UWIs in core analysis database
        found=0         #reset found flag
        for CUWI in cas_list:
            casName=CUWI[:-4]
            if UWI==casName:
                filename=CUWI
                found=1
        if found==0:                    #Create a new casfile
            filename=UWI+'.las'
            las_data=las.get_lasfile(filename,0)
            if not las_data:                     #canceled
                return
            #store EKB value
            if 'EKB' not in las_data.params:   #If no KB value look for alternative
                if 'EREF' not in las_data.params:
                    EKB=''
                else:
                    EKB=las_data.params['EREF'].value
            else:
                EKB=las_data.params['EKB'].value

            #create empty cas file
            mpath=global_vars.project.inDir+'/'+filename
            cas_data=tools.create_cas(mpath)
            if not cas_data:   #las creation failed
                return
            cas_data.params['GEO'] = LS.HeaderItem('GEO', value="", descr="Core examining geologist")
            cas_data.params['EKB'] = LS.HeaderItem('EKB', unit='M', value=EKB, descr='KB Elevation')
            cas_data.params['DSD'] = LS.HeaderItem('DSD', value="", descr='Date of core description')
            cas_data.params['CFL'] = LS.HeaderItem('CFL', value="", descr='Coring fluid')
            cas_data.params['CSH'] = LS.HeaderItem('CSH', unit='M',value="", descr='Core depth shift')

            #Create Dept curve
            #Find core in ws  - active worksheet
            ctop=0
            cbase=0
            cstep=round(las_data.well['STEP'].value,4)
            for mrow in ws:                    #Step through rows get top and base of core in UWI
                if mrow[0].value==UWI:
                    if ctop==0:
                        ctop=mrow[1].value
                    cbase=mrow[1].value + mrow[2].value

            mlist=[]                 #create empty list
            mlist.append([])         #create empty dept curve
            dept=ctop
            while dept<=cbase:
                mlist[0].append(dept)
                dept =round(dept+cstep,4)

            list_len=len(mlist[0])
            # Add empty core description and analysis curves
            mend=len(global_vars.cdescrvs)              #global list of core curve names

            # use null when converting to dataframe

            mynull=round(cas_data.well.NULL.value,2)
            mlist.append([])
            for c_idx in range(1,mend):   #depth curve is already created
                for l_idx in range(list_len):
                    mlist[c_idx].append(mynull)
                if c_idx<mend-1:
                    mlist.append([])

            #create cas-data with curves
            #Create CAS or Core ASCII file
            #First create the DEPT curve
            for c_idx in range(mend):
                #mnemonic=cdescrvs[c_idx]
                cas_data.insert_curve(ix=c_idx, mnemonic=global_vars.cdescrvs[c_idx], data=np.array(mlist[c_idx]),
                            unit= global_vars.c_units[c_idx], descr=global_vars.c_descr[c_idx],
                            value='')

        else:                            # Load cas file

            mend=len(global_vars.cdescrvs)              #global list of core curve names
            #load cas-data
            try:
                path=global_vars.project.inDir+'/'+filename
                cas_data = global_vars.LASCache.GetLASData(path)

            except FileNotFoundError:
                alert.Error('CAS file could not be found. Please, try again.')
                return
            except ValueError:
                alert.Error('CAS file could not be opened. Please, try again.')
                return
            except Exception:
                alert.Error('CAS file could not be read. Please, try again.')
                return
            '''
            #Check if core analysis already imported
            if len(cas_data.curves)>35:
                #Coredata already loaded go to next UWI
                c_Message(f"Core data already in {UWI}. To replace, delete cas-file in {Indir}.")
                continue
            '''
            #Prepare core data to insert into mlist core curves
            ctop=cas_data.well.STRT.value
            cbase=cas_data.well.STOP.value
            cstep=round(cas_data.well.STEP.value,1)
            mynull=cas_data.well.NULL.value

            #Create empty core curves in 'cas' file
            mlist=[]                 #create empty list
            mlist.append([])         #create empty dept curve
            mlist.append([])         #create dummy curve
            dept=ctop
            while dept<=cbase:
                mlist[0].append(dept)
                mlist[1].append(mynull)
                dept =round(dept+cstep,4)
            if len(cas_data.curves)<36:     #If core description cas exists
                for c_idx in range(35,mend):
                    #mnemonic=cdescrvs[c_idx]
                    cas_data.insert_curve(ix=c_idx, mnemonic=global_vars.cdescrvs[c_idx], data=np.array(mlist[1]),
                                unit= global_vars.c_units[c_idx], descr=global_vars.c_descr[c_idx],
                                value='')

        global_vars.core_df=pd.DataFrame(None)   #clear from previous well
        global_vars.core_df=cas_data.df()    #convert cas_data into data frame

        #find first core data point in core data base
        old_dpt=0
        for m_row in ws:                    #Step through rows get top and base of core in UWI
            if m_row[0].value==UWI:     #Find the well in the worksheet
                #Get the core data of the first row
                #'CORNO','S_TP','S_TCK','KMAX','KVRT','K90',
                #'CPOR','GDEN','BDEN','RSO','RSW'
                #find the core depth of the sample
                # Set index for mlists (to be substituted into cas_data)
                dept=m_row[1].value + m_row[2].value/2
                dept=round(ctop + int((dept-ctop)/cstep)*cstep,2)  #round to nearest depth increment
                if dept==old_dpt:      #If same as in previous sample then add one step
                    dept=dept+cstep
                    dept=round(ctop + int((dept-ctop)/cstep)*cstep,2)  #round to nearest depth increment
                old_dpt=dept

                #update core_df columns
                for c_dpt in global_vars.core_df.index:    #Cycle through dataframe
                    if  c_dpt>=dept and c_dpt<dept+cstep :
                        c_idx=0
                        global_vars.core_df.loc[c_dpt,'S_TP']=m_row[1].value
                        global_vars.core_df.loc[c_dpt,'S_TCK']=m_row[2].value
                        global_vars.core_df.loc[c_dpt,'CORNO']=m_row[3].value
                        for cname in global_vars.cdescrvs:
                            #skip core description curves and 'S_TP' and 'S_TCK'
                            if cname!='S_TP' and cname!='S_TCK' and cname!='CORNO':                #Do not overwrite 'S_TP' nor 'S_TCK'
                                if c_idx>=35:
                                    if not (m_row[c_idx-31].value):
                                        m_row[c_idx-31].value=round(mynull,4)
                                    if c_idx==39 or c_idx==40:                   #if density curves
                                        if m_row[c_idx-31].value>0:               #if NOT null
                                            global_vars.core_df.loc[c_dpt,cname]=(m_row[c_idx-31].value)/1000
                                        else:
                                            global_vars.core_df.loc[c_dpt,cname]=(m_row[c_idx-31].value)     #set null
                                    else:
                                        global_vars.core_df.loc[c_dpt,cname]=m_row[c_idx-31].value
                                c_idx +=1   # next column
                        break

        # Update cas_data
        #mynull=cas_data.well.NULL.value
        #core_df.fillna(mynull,inplace=True)
        #df=core_df.copy(deep=True)
        #cas_data.set_data(core_df)
        #cas_data.other=''

        # ready to save cas file
        cas_data.set_data(global_vars.core_df)    #update cas_data
        filename=UWI+'.cas'
        las.save_cas(filename,cas_data, Dir.In)
#==============================================================================================================
def c_zones(opt=0):
    '''
    Define/Edit Zones and AKAs (stratigraphic alternatives) using strat_col.xlsx
    Define/Edit Parameter file (opt 1)

    '''
    #Make sure Indir has been selected
    if not global_vars.project.inDir:
        alert.Error("Please, first select the Input directory")
        return

    if opt==0:
        mpath='databases/strat_col.xlsx'
    else:
        mpath='databases/Params.xlsx'

    editor : ExcelEditor = ExcelEditor(global_vars.project.inDir+'/'+mpath)
    editor.Show()

#PROJECT MANAGEMENT
def c_project():
    '''
    Create a project
    '''

    curDir=os.getcwd()
    os.chdir(global_vars.rootDir)

    global_vars.ui.Root.window.iconify()

    # Set prj_win toplevel window
    prj_win = common.initCommonWindow(title= 'Create New Project', width=550, height=300, topMost=2 )

    # get project name
    P_label0=Label(prj_win,text="Enter Project Name:")
    var1=StringVar()
    var1.set('project name')
    P_entry=ttk.Entry(prj_win,textvariable=var1)
    vin=StringVar()
    vin.set('Input (processed) Directory: Press <<Select>> for directory ')
    P_label1=Label(prj_win,textvariable=vin)
    vout=StringVar()
    vout.set('Output Directory: ')
    P_label2=Label(prj_win,textvariable=vout)
    vraw=StringVar()
    vraw.set('Unprocessed(Raw) LAS Directory: ')
    P_label3=Label(prj_win,textvariable=vraw)
    vcor=StringVar()
    vcor.set('Core data Directory: ')
    P_label4=Label(prj_win,textvariable=vcor)
    vclnt=StringVar()
    vclnt.set('Client LAS Directory: ')
    P_label5=Label(prj_win,textvariable=vclnt)

    #place widgets
    P_entry.grid(row=1,column=0, pady=5)
    P_label0.grid(row=0, column=0,columnspan=2, pady=5,sticky=W)
    P_label1.grid(row=2, column=0,pady=5,sticky=W)
    P_label2.grid(row=3, column=0,pady=5,sticky=W)
    P_label3.grid(row=4, column=0,pady=5,sticky=W)
    P_label4.grid(row=5, column=0,pady=5,sticky=W)
    P_label5.grid(row=6, column=0,pady=5,sticky=W)
    # done
    b_done=Button(prj_win, bg='cyan', text="Submit", command=lambda:p_submit(prj_win, var1.get(), 1))
    b_select=Button(prj_win, bg='orange', text='Select Directories', command= lambda:p_select(vin,vout,vraw,vcor,vclnt))
    b_cancel=Button(prj_win, bg='pink', text='Cancel', command= lambda:p_submit(prj_win,0, 0))

    b_done.grid(row=8,column=0, padx=20,pady=10,sticky=W)
    b_select.grid(row=8,column=0,padx=40, pady=10)
    b_cancel.grid(row=8,column=0,padx=20,pady=10,sticky=E)

    prj_win.mainloop()
    global_vars.ui.Root.window.deiconify()
    os.chdir(curDir)
#-----------------------------------------------------------------------------------------------------
def p_select(vin,vout,vraw,vcor, vclnt):
    # get Indir
    prompt.InputDir()
    vin.set('Input (processed) Directory: ...' + global_vars.project.inDir[-40:])
    # get Outdir
    prompt.OutputDir()
    vout.set('Output Directory: ...' + global_vars.project.outDir[-40:])
    # get Rawdir
    prompt.RawDir()
    vraw.set('Unprocessed(Raw) LAS Directory: ...' + global_vars.project.rawDir[-40:])
    # get Coredir
    prompt.CoreDir()
    vcor.set('Core data Directory: ...' + global_vars.project.coreDir[-40:])
    # get Clientdir
    prompt.ClientDir()
    vclnt.set('Client Directory: ...' + global_vars.project.clientDir[-40:])
#-----------------------------------------------------------------------------------------------------
def p_submit(prj_win, var1, opt):
    '''
    Submit project parms and store in Project list
    '''

    if opt==1:      # If project data submitted
        global_vars.project.settings.Set('name', var1)
        global_vars.project.Save()
        global_vars.project.SaveInit()

        global_vars.ui.Root.window.deiconify()
        global_vars.ui.Root.Update()

    #close
    main.stat_update('dummy', 6)                                                       # Reload previous project????
    prj_win.destroy()
# ====================================================================================================
# Define root window
def main_func():
    global_vars.running = True
    global_vars.login = Login()
    if not global_vars.login.loggedIn:
        return
    
    #init and load ui
    global_vars.ui = classes.UI()
    global_vars.ui.Load()
    
    #auto_loadproj()
    global_vars.project = Project(global_vars.rootDir+'/config/init.crv')
    global_vars.project.Load()

    global_vars.ui.Root.Show()

    print("end of main - close")
    
if __name__ == "__main__":
    main_func()